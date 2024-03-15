import sys
import getopt
import os
import openpyxl


def main(argv):
    template_file = '{ВПXXXX} {Име на линията} - проверки.xlsx'

    if not os.path.isfile(template_file):
        print(f'Template file {template_file} not found!')
        sys.exit()

    input_file = process_arguments(argv)
    output_file = get_output_file_name(input_file)

    wb_source = openpyxl.load_workbook(input_file)  # Създаване на нов Excel файл - входния
    wb_target = openpyxl.load_workbook(template_file)  # Отваря темплейтния файл

    copy_data(wb_source, wb_target, 'Стълб', ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'])
    copy_data(wb_source, wb_target, 'Оборудване на стълб', ['A', 'B', 'C', 'D', 'E'])
    copy_data(wb_source, wb_target, 'РОМ - РОС', ['A', 'B', 'C', 'D'])
    copy_data(wb_source, wb_target, 'Вентилен отвод', ['A', 'B', 'C'])
    copy_data(wb_source, wb_target, 'ИЗКС', ['A', 'B'])
    copy_data(wb_source, wb_target, 'Проводник', ['A', 'B'])

    # 'Стълб.Подтип'
    transform_value_in_cell(wb_target, 'Стълб', 'C', map_pole_subtype)
    # 'Оборудване на стълб.Вид оборудване'
    transform_value_in_cell(wb_target, 'Оборудване на стълб', 'B', map_pole_equipment_type)
    # 'РОМ - РОС.Експлоатационно състояние'
    transform_value_in_cell(wb_target, 'РОМ - РОС', 'B', map_sectional_status)
    # 'РОМ - РОС.Вид'
    transform_value_in_cell(wb_target, 'РОМ - РОС', 'C', map_sectional_subtype)
    # 'Вентилен отвод.Фази'
    transform_value_in_cell(wb_target, 'Вентилен отвод', 'B', map_surge_arrestor_phase_designation)

    wb_target.save(output_file)
    print(f'File saved: {output_file}')


def process_arguments(argv):
    opts, args = getopt.getopt(argv, 'hi:')
    input_file = None

    for opt, arg in opts:
        if opt == '-h':
            print('CreateCommentsFile.py -i <inputfile>')
            sys.exit()
        elif opt == '-i':
            input_file = arg

    if not os.path.isfile(input_file):
        print(f'Input file {input_file} not found!')
        sys.exit()

    return os.path.abspath(input_file)


def get_output_file_name(input_file):
    output_directory = os.path.dirname(input_file)
    vals = os.path.basename(input_file).split('-')

    return f"{output_directory}\\{vals[0]} {vals[2]} - comments.xlsx"


def delete_column(wb, sheet, col):
    ws = wb[sheet]
    ws.delete_cols(col)


def copy_data(wb_source, wb_target, sheet, cols):
    ws_source = wb_source[sheet]
    ws_target = wb_target[sheet]

    for col in cols:
        for row, cell in enumerate(ws_source[col], start=2):  # start=2 защото започваме от ред 2
            ws_target.cell(row=row, column=cell.column, value=cell.value)


def transform_value_in_cell(wb, sheet, col, func):
    ws = wb[sheet]
    cells = ws[col]
    for cell in cells:
        cell.value = func(cell.value)


def map_pole_subtype(key):
    subtype_mapping = {
        '0': 'Стълб - недиференцирано',
        '1': 'Дървен',
        '2': 'Композитен',
        '3': 'Стоманорешетъчен',
        '4': 'Стоманотръбен',
        '5': 'Стоманобетонен'}

    return subtype_mapping.get(key, 'N/A')


def map_pole_equipment_type(key):
    equipment_mapping = {
        '0': 'Оборудване на стълб',
        '1': 'Изолатор',
        '2': 'Конзола',
        '3': 'Обтяжка/подпора',
        '4': 'Заземителен контур'}

    return equipment_mapping.get(key, 'N/A')


def map_sectional_status(key):
    status_mapping = {
        '0': 'Изключено',
        '1': 'Включено',
        '2': 'НЯМА ИНФОРМАЦИЯ'}

    return status_mapping.get(key, 'N/A')


def map_sectional_subtype(key):
    subtype_mapping = {
        '0': 'Секциониращ разединител',
        '1': 'Товаров разединител с ДУ',
        '2': 'Реклоузер',
        '3': 'РОМ',
        '4': 'РОС'}

    return subtype_mapping.get(key, 'N/A')


def map_surge_arrestor_phase_designation(key):
    phase_mapping = {
        '0': 'НЯМА ИНФОРМАЦИЯ',
        '1': 'L3',
        '2': 'L2',
        '3': 'L2-L3',
        '4': 'L1',
        '5': 'L1-L3',
        '6': 'L1-L2',
        '7': 'L1-L2-L3'}

    return phase_mapping[key]   #get връща стойността за даден ключ ; key се използва веднъж за търсене в речника ; и ако ключът key не е намерен в речника, се връща самата стойност на key
    #return phase_mapping.get(key, key)


if __name__ == "__main__":
    main(sys.argv[1:])
